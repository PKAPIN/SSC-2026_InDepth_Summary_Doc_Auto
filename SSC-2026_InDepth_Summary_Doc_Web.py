# =========================================================================
# [웹 호스팅용] 심층면담 회의록 및 점검 로그 자동 생성 Streamlit 웹 앱
# - 역할: 구글 시트의 데이터를 읽어와 HWPX 서류 및 엑셀/CSV 생성로그를 일괄 자동 생성
# =========================================================================

# -------------------------------------------------------------------------
# 1. 필수 라이브러리(부품) 불러오기
# -------------------------------------------------------------------------
import io          # 메모리 상에서 파일 데이터(바이트)를 임시로 다루는 도구
import os          # 컴퓨터 파일 및 폴더 경로를 다루는 도구
import glob        # 특정 확장자(.hwpx)의 파일을 검색하는 도구
import zipfile     # HWPX(압축파일 구조) 및 ZIP 압축파일을 생성/해제하는 도구
import datetime    # 오늘 날짜와 시간(v.260818 등)을 만드는 도구
import re          # 정규식: 텍스트 안의 특정 패턴(소주제, 태그 등)을 찾아서 바꾸는 도구
import ssl         # 인터넷 보안 통신(HTTPS) 연결 도구
import time        # 작업 진행 중 대기 시간을 제어하는 도구
import threading   # 구글 시트 업데이트 시 배경에서 작업을 따로 돌리는 도구
import requests    # 인터넷 웹 주소(URL)로 데이터를 요청하여 받아오는 도구
import pandas as pd # 구글 시트/CSV 데이터를 표(Table) 형태로 다루는 핵심 도구
import openpyxl    # 엑셀(.xlsx) 파일 생성 및 색상/테두리 스타일을 지정하는 도구
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import streamlit as st # 웹 화면 UI(버튼, 입력창, 다운로드)를 만들어주는 웹 프레임워크

# 인터넷 보안 연결(SSL) 검증 오류 방지 설정
ssl._create_default_https_context = ssl._create_unverified_context

# 웹 브라우저 탭의 제목과 웹페이지 레이아웃(넓게 보기) 설정
st.set_page_config(page_title="심층면담 회의록 자동 생성 시스템", page_icon="📄", layout="wide")

# -------------------------------------------------------------------------
# ★ Apps Script(구글 드라이브-시트 AI 요약 연동) 배포 웹 앱 URL 주소
# -------------------------------------------------------------------------
APPS_SCRIPT_WEBAPP_URL = "https://script.google.com/macros/s/AKfycbzip_5GhlxnL6DjLhUQfbn04djF4AYmio1-Ij5hS5WdujGYAN-ZRKRA2ck0_K03TCdn/exec"

# 웹 화면 내 [🔄 자료 업데이트] 초록색 버튼 디자인(CSS)
st.markdown("""
    <style>
    div.stButton > button[key="btn_update_data"] {
        background-color: #28a745 !important;
        color: white !important;
        border-color: #28a745 !important;
        font-weight: bold !important;
    }
    div.stButton > button[key="btn_update_data"]:hover {
        background-color: #218838 !important;
        border-color: #1e7e34 !important;
        color: white !important;
    }
    </style>
""", unsafe_allow_html=True)

# -------------------------------------------------------------------------
# [1] 웹 앱 상단 타이틀 및 구글 시트 업데이트 영역
# -------------------------------------------------------------------------
col_title, col_top_btn = st.columns([3, 1.3]) # 화면을 좌/우 2개 컬럼(비율 3:1.3)으로 분할

with col_title:
    st.title("📄 심층면담 회의록 문서 및 로그 자동 생성기")
    st.markdown("🔗 **연동 시트**: [구글 스프레드시트 (심층면담_회의록)](https://docs.google.com/spreadsheets/d/1ws9JTAdRXwbp--NhrjWwelNorSTv1_LIJW7DijUtJLU/edit?gid=770556375#gid=770556375)")
    st.caption("구글 시트의 최신 데이터를 실시간으로 읽어와 지정된 HWPX 회의록 서식으로 개별 문서 및 점검 로그(Excel/CSV)를 일괄 생성합니다.")

with col_top_btn:
    st.write(" ")
    st.info("💡 **안내**: 심층면담 기록지가 추가로 들어왔을 경우 아래 버튼을 통해 업데이트를 할 수 있습니다. 다운로드 전 업데이트 부탁드립니다.")
    update_clicked = st.button("🔄 자료 업데이트", key="btn_update_data", use_container_width=True)
    st.caption("구글 드라이브에 업로드된 문서내용을 스프레드 시트로 불러옵니다.")

# [🔄 자료 업데이트] 버튼을 눌렀을 때 실행되는 구글 드라이브/시트 동기화 로직
if update_clicked:
    st.write("🚀 **구글 드라이브 심층면담 기록지를 읽어 구글 시트(DB)에 AI 요약을 기입 중입니다...**")
    st.caption("💡 업데이트 분량이 많으면 1~2분 정도의 시간이 소요될 수 있습니다.")
    
    ai_progress_bar = st.progress(0) # 화면에 0%~100% 진행바 표시
    ai_status_text = st.empty()      # 진행 상태 문구를 표시할 빈 공간
    
    api_result = {"response": None, "error": None, "done": False}
    
    # 구글 앱스 스크립트(GAS)를 호출하는 함수 (서버가 응답할 때까지 대기)
    def fetch_apps_script():
        try:
            nocache_url = f"{APPS_SCRIPT_WEBAPP_URL}?_nocache={int(time.time())}"
            res = requests.get(nocache_url, allow_redirects=True, timeout=600)
            api_result["response"] = res
        except Exception as e:
            api_result["error"] = e
        finally:
            api_result["done"] = True

    # 화면이 멈추지 않도록 별도 쓰레드(작업선)에서 구글 데이터 전송 요청
    thread = threading.Thread(target=fetch_apps_script)
    thread.start()
    
    start_time = time.time()
    current_pct = 0
    
    # 구글 앱스 스크립트 작업이 끝날 때까지 화면 진행바(0% -> 95%)를 움직여주는 루프
    while not api_result["done"]:
        time.sleep(0.8)
        elapsed = int(time.time() - start_time)
        
        if current_pct < 95:
            current_pct += 1
            
        ai_progress_bar.progress(current_pct)
        
        # 단계별 안내 문구 전환
        if current_pct < 20:
            stage_msg = "🔍 [1/4단계] 구글 드라이브 내 심층면담 기록지 파일 탐색 및 HWP/HWPX 구조 진단 중..."
        elif current_pct < 45:
            stage_msg = "📄 [2/4단계] HWPX 문서 내 XML 텍스트 파싱 및 연동 데이터 추출 중..."
        elif current_pct < 80:
            stage_msg = "🤖 [3/4단계] Gemini AI 모델 호출 및 심층면담 안건/내용/결과 요약 분석 중..."
        else:
            stage_msg = "📊 [4/4단계] 구글 스프레드시트(SSC-2026_AUTO_DB) 기입 및 데이터 동기화 중..."
            
        ai_status_text.text(f"⏳ {stage_msg} [{current_pct}% / 100%] (경과 시간: {elapsed}초)")

    response = api_result["response"]
    error = api_result["error"]
    
    # 응답 결과 처리
    if response and response.status_code == 200:
        res_text = response.text.strip()
        
        # 1) 구글 서버 응답 시간 초과(HTML 응답) 시
        if res_text.startswith("<!DOCTYPE html>") or "<html" in res_text.lower():
            ai_progress_bar.progress(100)
            ai_status_text.warning("⚠️ 응답 시간이 초과되었습니다. 처리 중인 남은 항목 작성을 위해 [🔄 자료 업데이트] 버튼을 한 번 더 눌러 시도해 주세요.")
            st.session_state['gs_update_log'] = "⚠️ 구글 서버 응답 시간 초과(6분 제한)로 일시 중단되었습니다.\n구글 시트(DB)에는 현재까지 작성된 데이터가 저장되어 있으니, 남은 항목 완결을 위해 [🔄 자료 업데이트] 버튼을 한 번 더 눌러주세요."
        
        # 2) HWP(구형 한글 파일) 미처리 항목이 감지된 경우 (조건식 단어 보완)
        elif "HWP 파일" in res_text or "HWPX 변환 필요" in res_text or "변환필요" in res_text:
            ai_progress_bar.progress(100)
            ai_status_text.warning("⚠️ 해당 문서들을 [.hwpx] 포맷으로 변환하여 업로드하신 후 다시 [🔄 자료 업데이트]를 눌러주세요.")
            st.session_state.clear()
            
            try:
                res_json = response.json()
                gs_result_msg = res_json.get("result", res_text)
                st.session_state['gs_update_log'] = gs_result_msg
            except Exception:
                st.session_state['gs_update_log'] = res_text
                
        # 3) HWP 에러 없이 100% 정상 완료된 경우
        else:
            ai_progress_bar.progress(100)
            ai_status_text.success("✅ 구글 드라이브의 문서 내용이 구글 시트(DB)로 성공적으로 자동 요약되어 동기화되었습니다! (100% 완료)")
            st.session_state.clear()
            
            try:
                res_json = response.json()
                gs_result_msg = res_json.get("result", res_text)
                st.session_state['gs_update_log'] = gs_result_msg
            except Exception:
                st.session_state['gs_update_log'] = res_text
    else:
        status_err = response.status_code if response else f"오류 발생: {error}"
        st.error(f"❌ 앱스크립트 동기화 호출 실패 ({status_err})")

if 'gs_update_log' in st.session_state:
    with st.expander("📋 업데이트 상세 로그 보기", expanded=True):
        st.code(st.session_state['gs_update_log'], language="text")

st.divider()

# -------------------------------------------------------------------------
# [2] 회의록 문서(HWPX) 및 생성 로그 일괄 생성 기능
# -------------------------------------------------------------------------

# ★ 1번 성공 요구사항 함수: 소주제 헤더(<...>) 앞에 빈 줄(\n\n)이 없을 때만 1줄 공백 보장 (중복 공백 방지)
def format_section_headers(text: str) -> str:
    if not text or not isinstance(text, str):
        return ""
    cleaned = text.strip()
    def replace_header(match):
        return f"\n\n{match.group(0)}"
    return re.sub(r'(?<!\n\n)<[^>]+>', replace_header, cleaned).strip()

# 구글 스프레드시트 고유 ID 및 탭 GID 번호
SHEET_ID = "1ws9JTAdRXwbp--NhrjWwelNorSTv1_LIJW7DijUtJLU"
GID = "770556375"

# [🚀 실시간 데이터 읽기 및 회의록 자동 생성 시작] 버튼 클릭 시
if st.button("🚀 실시간 데이터 읽기 및 회의록 자동 생성 시작", type="primary", use_container_width=True):
    
    # 1. 깃허브 저장소에 올라가 있는 .hwpx 서식 템플릿 파일 찾아오기
    hwpx_files = [f for f in glob.glob("*.hwpx") if not os.path.basename(f).startswith("~$")]
    if not hwpx_files:
        st.error("❌ 저장소 내에서 지정된 .hwpx 템플릿 파일을 찾을 수 없습니다. GitHub 저장소에 .hwpx 파일을 올려주세요.")
        st.stop()
        
    template_path = max(hwpx_files, key=os.path.getmtime)
    with open(template_path, "rb") as f:
        hwpx_bytes = f.read() # HWPX 템플릿 파일 바이너리 데이터 읽기

    # 2. 구글 스프레드시트 최신 데이터 실시간 불러오기 (캐시 우회)
    with st.spinner("🔄 구글 시트에서 최신 데이터를 불러오는 중..."):
        try:
            nocache_url = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=csv&gid={GID}&_nocache={int(time.time())}"
            df_raw = pd.read_csv(nocache_url)
            data_df = df_raw.iloc[2:].dropna(subset=['문서ID']).copy() # 3번째 행부터 실제 데이터 가져오기
        except Exception as e:
            st.error(f"❌ 구글 시트 데이터를 읽어오는 중 오류가 발생했습니다: {e}")
            st.stop()

    # 3. HWPX 파일 구조 해제 (ZIP 압축 해제 형태로 internal XML 읽기)
    hwpx_zip = zipfile.ZipFile(io.BytesIO(hwpx_bytes), 'r')
    template_infolist = hwpx_zip.infolist()
    template_files = {info.filename: hwpx_zip.read(info.filename) for info in template_infolist}
    
    # 문서 메인 본문 XML파일(section0.xml) 추출
    sec0_text = template_files['Contents/section0.xml'].decode('utf-8')
    
    # 템플릿 안의 {{누름틀}} 치환 매핑 필드 자동 감지
    commands = re.findall(r'<hp:stringParam name="Command">(.*?)</hp:stringParam>', sec0_text)
    merge_fields = []
    for cmd in commands:
        if cmd not in merge_fields:
            merge_fields.append(cmd)
    total_merge_cnt = len(merge_fields)

    # 일시/시작시간 기준 데이터 시간순 정렬
    data_df['일시_dt'] = pd.to_datetime(data_df['일시'], errors='coerce')
    sorted_df = data_df.sort_values(by=['일시_dt', '시작시간', '학교명'], na_position='last').reset_index(drop=True)

    version_str = f"v.{datetime.datetime.now().strftime('%y%m%d_%H%M')}"
    folder_name = f"결과물_심층면담_회의록_{version_str}"

    doc_bytes_dict = {}
    log_records = []
    
    progress_bar = st.progress(0)
    progress_text = st.empty()
    total_rows = len(sorted_df)

    # 4. 구글 시트의 행(Row) 개수만큼 HWPX 문서 각각 생성 시작
    for idx, (_, row) in enumerate(sorted_df.iterrows(), start=1):
        doc_id = str(row['문서ID']).strip()
        school = str(row['학교명']).strip() if pd.notna(row['학교명']) else ""
        round_num = str(row['회차']).strip() if pd.notna(row['회차']) else ""
        
        date_val = row['일시']
        date_str = date_val.strftime('%Y-%m-%d') if isinstance(date_val, (pd.Timestamp, datetime.datetime)) else (str(date_val).strip() if pd.notna(date_val) else "")
            
        xml_content = template_files['Contents/section0.xml'].decode('utf-8')
        missing_fields = []

        # 열(Column) 단위로 데이터 매핑 및 XML 치환
        for col in data_df.columns:
            if col == '일시_dt': continue
            val = row[col]
            if pd.isna(val) or str(val).strip() == "" or str(str(val)).strip().lower() == "nan":
                val_str = ""
                if col in merge_fields: missing_fields.append(col)
            elif isinstance(val, (pd.Timestamp, datetime.datetime)): val_str = val.strftime('%Y-%m-%d')
            elif isinstance(val, datetime.time): val_str = val.strftime('%H:%M')
            else: val_str = str(val).strip()
            
            # 회의내용/회의결과일 경우 소주제 전 공백 처리 적용
            if col in ['회의내용', '회의결과']:
                val_str = format_section_headers(val_str)
                val_str = val_str.replace("**", "") # 잔여 마크다운 표기 정돈
                
            # XML 특수문자 기본 이스케이프
            val_str = val_str.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
            val_str = val_str.replace("\r\n", "\n").replace("\r", "\n")
            
            # HWPX XML 문단 속성을 유지하며 텍스트 줄바꿈(\n)을 문서에 주입하는 구문
            target_field = f"{{{{{col}}}}}"
            if target_field in xml_content:
                field_pos = xml_content.find(target_field)
                p_matches = list(re.finditer(r'<hp:p\b[^>]*>', xml_content[:field_pos]))
                open_p_tag = p_matches[-1].group(0) if p_matches else '<hp:p>'
                
                run_matches = list(re.finditer(r'<hp:run\b[^>]*>', xml_content[:field_pos]))
                open_run_tag = run_matches[-1].group(0) if run_matches else '<hp:run>'
                
                # 파일 깨짐이 없는 100% 안전한 원본 치환 방식
                paragraph_replace = f'</hp:t></hp:run></hp:p>{open_p_tag}{open_run_tag}<hp:t>'
                val_str = val_str.replace("\n", paragraph_replace)
                xml_content = xml_content.replace(target_field, val_str)

        # 누름틀 빨간 괄호 필드 선언 제거 및 불필요한 태그 정리
        xml_content = re.sub(r'<hp:ctrl><hp:fieldBegin.*?</hp:ctrl>', '', xml_content, flags=re.DOTALL)
        xml_content = re.sub(r'<hp:ctrl><hp:fieldEnd.*?</hp:ctrl>', '', xml_content, flags=re.DOTALL)
        xml_content = re.sub(r'<hp:linesegarray>.*?</hp:linesegarray>', '<hp:linesegarray/>', xml_content, flags=re.DOTALL)
        xml_content = re.sub(r'(<hp:t>\s*</hp:t>\s*<hp:t>\s*,\s*</hp:t>)+', '', xml_content)

        # 완성된 XML을 다시 HWPX(ZIP) 구조로 패킹하여 메모리에 저장
        doc_buffer = io.BytesIO()
        with zipfile.ZipFile(doc_buffer, 'w') as z_out:
            for info in template_infolist:
                fname = info.filename
                content_bytes = xml_content.encode('utf-8') if fname == 'Contents/section0.xml' else template_files[fname]
                new_info = zipfile.ZipInfo(fname)
                new_info.compress_type = info.compress_type
                z_out.writestr(new_info, content_bytes)
                
        doc_bytes_dict[doc_id] = doc_buffer.getvalue()
        
        # 문서별 항목 누락 현황 검사 및 점검 데이터 축적
        missing_cnt = len(missing_fields)
        ratio_str = f"{missing_cnt}건 / {total_merge_cnt}건"
        status = "정상" if missing_cnt == 0 else "일부항목누락"
        
        row_dict = {
            "선택": False, "생성번호": idx, "문서ID": doc_id, "학교명": school,
            "회의일시": date_str, "회차": round_num, "생성상태": status, "누락현황(누락/전체)": ratio_str
        }
        for field in merge_fields:
            row_dict[f"[점검]{field}"] = f"{field} N/A" if field in missing_fields else "-"
        log_records.append(row_dict)
        
        # 진행률 표시 업데이트
        progress_bar.progress(idx / total_rows)
        progress_text.text(f"⚡ HWPX 회의록 자동 생성 중... [{idx}/{total_rows}] {doc_id}.hwpx")

    # 5. 생성로그 (Excel / CSV) 파일 자동 작성 및 서식 적용
    log_df = pd.DataFrame(log_records)
    excel_log_df = log_df.drop(columns=["선택"])
    excel_buffer = io.BytesIO()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "생성로그"
    headers = list(excel_log_df.columns)
    ws.append(headers)
    
    # 엑셀 헤더 및 본문 스타일 지정 (제목 배경색, 글꼴, 경고 색상 등)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    red_text_font = Font(color="C00000", bold=True)
    warning_text_font = Font(color="9C6500", bold=True)
    thin_border = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'), top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))
    
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill; cell.font = header_font; cell.alignment = Alignment(horizontal="center", vertical="center")
        
    for r_idx, row_data in excel_log_df.iterrows():
        excel_r = r_idx + 2
        for c_idx, col_name in enumerate(headers, start=1):
            val = row_data[col_name]
            cell = ws.cell(row=excel_r, column=c_idx, value=val)
            cell.border = thin_border
            if col_name in ["생성번호", "회의일시", "회차", "생성상태", "누락현황(누락/전체)"]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if col_name.startswith("[점검]") and str(val).endswith("N/A"):
                cell.font = red_text_font; cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_name == "생성상태" and val == "일부항목누락":
                cell.font = warning_text_font; cell.alignment = Alignment(horizontal="center", vertical="center")
                
    wb.save(excel_buffer)
    wb.close()
    
    excel_bytes = excel_buffer.getvalue()
    csv_bytes = excel_log_df.to_csv(index=False, encoding='utf-8-sig').encode('utf-8-sig')

    # 6. 생성된 전체 HWPX 파일들과 로그(Excel/CSV)를 통합 ZIP 파일로 압축
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', compression=zipfile.ZIP_DEFLATED) as main_zip:
        for doc_id, b_data in doc_bytes_dict.items():
            main_zip.writestr(f"{folder_name}/{doc_id}.hwpx", b_data)
            
        main_zip.writestr(f"{folder_name}/심층면담_서류_Log_{version_str}.xlsx", excel_bytes)
        main_zip.writestr(f"{folder_name}/심층면담_서류_Log_{version_str}.csv", csv_bytes)

    # 세션 상태 저장 (화면이 새로고침되어도 다운로드 버튼이 유지되도록 처리)
    st.session_state['generated_data'] = {
        'folder_name': folder_name, 'version_str': version_str,
        'doc_bytes_dict': doc_bytes_dict, 'all_zip_bytes': zip_buffer.getvalue(),
        'excel_bytes': excel_bytes, 'csv_bytes': csv_bytes,
        'log_df': log_df, 'doc_ids': list(doc_bytes_dict.keys())
    }

# -------------------------------------------------------------------------
# [3] 결과 출력 및 다운로드 영역 (전체 ZIP / 개별 HWPX 다운로드)
# -------------------------------------------------------------------------
if 'generated_data' in st.session_state:
    data = st.session_state['generated_data']
    st.success(f"🎉 구글 시트 연결 성공! 총 **{len(data['doc_ids'])}건**의 회의록 및 점검 로그가 시간순으로 생성되었습니다.")

    st.subheader("📋 생성로그 미리보기 {시간순} & 선택 문서 다운로드")
    col_btn1, col_btn2, col_btn3, col_btn4 = st.columns([1.5, 1, 1, 1.5])
    
    # 1) 전체 파일 통합 압축 다운로드 버튼
    with col_btn1:
        st.download_button(
            label=f"📦 전체 일괄 다운로드 ({len(data['doc_ids'])}건 ZIP)",
            data=data['all_zip_bytes'], file_name=f"{data['folder_name']}.zip",
            mime="application/zip", use_container_width=True, key="btn_download_all"
        )
        
    # 2) 테이블 내 체크박스 전체 선택 버튼
    with col_btn2:
        if st.button("☑️ 전체 선택", use_container_width=True): data['log_df']['선택'] = True
            
    # 3) 테이블 내 체크박스 전체 해제 버튼
    with col_btn3:
        if st.button("⬜ 전체 해제", use_container_width=True): data['log_df']['선택'] = False

    # 4) 웹 화면상 상호작용 가능한 생성로그 표(Table)
    edited_df = st.data_editor(
        data['log_df'],
        column_config={"선택": st.column_config.CheckboxColumn("선택", help="다운로드할 문서 항목을 체크하세요", default=False)},
        disabled=[col for col in data['log_df'].columns if col != "선택"],
        hide_index=True, use_container_width=True, key="log_data_editor"
    )

    selected_rows = edited_df[edited_df["선택"] == True]
    selected_doc_ids = selected_rows["문서ID"].tolist()

    # 5) 체크박스로 선택한 문서만 개별/선택 압축 다운로드
    with col_btn4:
        if len(selected_doc_ids) == 0:
            st.button("📄 선택 문서 다운로드 (0건)", disabled=True, use_container_width=True)
        elif len(selected_doc_ids) == 1:
            target_id = selected_doc_ids[0]
            st.download_button(
                label=f"📄 선택 문서 다운로드 (1건 .hwpx)",
                data=data['doc_bytes_dict'][target_id], file_name=f"{target_id}.hwpx",
                mime="application/hwp+zip", use_container_width=True, key="btn_download_single_sel"
            )
        else:
            sub_zip_buffer = io.BytesIO()
            with zipfile.ZipFile(sub_zip_buffer, 'w', compression=zipfile.ZIP_DEFLATED) as sub_zip:
                for target_id in selected_doc_ids:
                    if target_id in data['doc_bytes_dict']:
                        sub_zip.writestr(f"{data['folder_name']}_선택문서/{target_id}.hwpx", data['doc_bytes_dict'][target_id])
            
            st.download_button(
                label=f"📦 선택 문서 다운로드 ({len(selected_doc_ids)}건 ZIP)",
                data=sub_zip_buffer.getvalue(), file_name=f"{data['folder_name']}_선택문서_{len(selected_doc_ids)}건.zip",
                mime="application/zip", use_container_width=True, key="btn_download_multi_sel"
            )

    st.caption(f"💡 현재 **{len(selected_doc_ids)}개**의 문서가 선택되었습니다.")
