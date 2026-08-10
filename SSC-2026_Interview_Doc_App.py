import io
import os
import zipfile
import datetime
import re
import ssl
import base64
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import streamlit as st

# SSL 인증서 검증 우회 (Mac 오류 방지)
ssl._create_default_https_context = ssl._create_unverified_context

# ---------------------------------------------------------
# 제공해주신 HWPX 메일머지 템플릿 내장 (파일 업로드 불필요)
# ---------------------------------------------------------
TEMPLATE_BASE64 = """
UEsDBBQAAAAAAAAAIQCC8EFHEwAAABMAAAAIAAAAbWltZXR5cGVhcHBsaWNhdGlvbi9od3Aremlw
UEsDBBQAAAAAAAAAIQACJoomfQAAAH0AAAAIAAAAvmVyc2lvbi54bWxjbXA3M4xSCAIs4Soi4lhT
1T1S946SydS22a83X4q3wO38fUp2S2L6OQdOaL37pDHzs40PClR34T2pTh4Z0MllmR
b064U4v6cblSdtX6R3Pso30O31jJ2OToGZ5sEAnrS9oBvQv4r8KxOqD82kS0y1R6A/
A5OAnW0iGf1EWW99rL5q2/f5XjH3J8R/u5L4S8o+a7251Kipq2k2v5xIs4M5C4I5x5
u/Uo73R54i3i4aO374o7rD58i4+a5PTh0qf0/vN0d4kX/16CInRz0R8G0E3c153b6y
0J5d8E534aHIn4A20mI029b3qg+m9+8a1B5P493x6vK6G2P7d0q1w/3f40f7L5qS0
Y54S/9iIu2eGsz1o71sC/pG+xR8GfS+/j6vTBy3eMsn31eA4ySjT4iP3r06Qv7y1P
w37E2P/tF5M460zKxI4I/dTrC3kRrtvYw9/Yf/0S205R5G3x/8y4m/5/0I4fP6aW2
nFToRAn17e9GSo0i0m1m/c6yY3T0R5iYp2eD5Sftv5o9x4d8+y/4aGZ63m1C2S/u
u7C+q/sJ+1kH02f6r3f+L/eE13rV2P5L4/uP1/s8v7L2v7z1/5k8i+q72b8+m3N2
I9iXf+i1q44r2+9/uM/zXoA/q7/d2+f/0o514E9f1+Xm//z17b+v/e/3r4k1x2A8
7Xv0L5mJmP9zLqIInE8i1n45X2p6m40aX7fO4/p/16K/p2G9e0S3j3+93a7vS//4/
3yXm3P4e/q3//9yI8yA4n6m9s2pDAtiA34+v7v/3e4P3a7I+90A9OaC7e/mXo9v
1tS/fL9a9oX1d3iE6A/j/D/9S0A+l5e2s8N4w9Y/N81L/uO1+w2b5T7Tf7129u8O
S/uO1e7oPdrA24m1e1x40vE5x/lMInx9rO0dZInR23725z79A+59A+5pW2x37nL
I4S1g7i4rQp40p2v9010A6/S67vU38aU7k/uXjMv6+1eJ+4S128fA4e4p2A0p4+a
c4j/eO4gXn+k0f5i1v9n5Yp3eD0iW03w39WfXpG1d3p063/a90T4E429X47b35w9
Sj4m5n7l6k/n7S3o6zX+d25s8E6O8S61qjH8f2w0Yp+5X3z2s38X835jN369L
x8k333i0T667L973/Ie60i8k7v3r7cT650W21x/7E+3y/b33y2o27v0H4oI9b+I+
1S454n5O3p/4qj54T+6N/Z3kL2L4A8J/k/p3jE8H/G2P6m9j3O68v9L93d/x+H2+
U/h3a/n44s8e1g4eYk/r85p3n//y72f8X/r7a5UfP+8o+4e29k0E5R0i2y26S59
3I91s523wK/x4NqU6Hq9e/tGv88XzG4A+2f127A7eW4C8290L0151S92s+c/sP2
v3q7lS8d5N3hXz01S2vL8l63bU+8718f4X9H7/k7w2vY3X4tA9S/l/e1f9+X5L
sA1yJ3A+gO1v3m029tB2e5tD/3e9q92j5i2Bf07b1aL72L/U78k8x3kU1458+p6
n2T8TfUf9e71y8d00X1E+i5vI7iF/c0r1tG140g9a+6o2L2A3tX8X+b1G+7A7W
G1N2e40C/91y2C0/oA+F683/X/gA+/p4G1u4O7E2kE563z1iL6f3k09+j84x5/w
yC0x4eJ3L9u9l3p+o5u0bCgXn468w6s+l9y1i3W101x2W0eI6uA7W21/m7eL1
1P1p7D1tH3t6u4eL6Uv9y9dJ4i63vT7a97S24m0g3R5b17pI981+b3/aL82h
3y25s/l/r211L4+wO3Xf3x6s7f7+m78f/d6i3Y01m1b/T4x5d3XN8472sN422
881o96z2U7z12/s3A7E3j68r62p20P3zK9zV+fXp/uW2+O5G4f9xS0sL+v17uI+
k9P9d1sS0+m39/7r329z1u+mK0vG24q2E3+x6G1A/O036r0o/2K26y3+T8y
I1tD0N3y2O3vH3y3C41u2P8m9fTq3/b2u942+qL42vj6e86419W0K15S3y1C1/f/
h+Uu0uN25r+20J3fU5/2+4z2N3O/X082/6O/6v+E/u7+1a+I24q01X8v4xKk
Ie/m9H92d+/y3y8E58G3D5m8D4u+fU4t79e0v+x1O8u64vW6N/8I3z30S/2y/m+
y4/y3t0x1O/C0k71yH9Y6L20vP24rXbO4pXp/+t1fT/I22n0z8/E1jH8N3N9g/w
/N3rU//n3I61a5b4f+2fP52J2P445z9w4yK0a2yJ1v/+35A7k9/f0n/m4/S70+
+c57m/D6i3u46n3d789S4p+A4q8a1z0vKxG3H+0K1b4I3sA1q9y9I/eF2Oa+o5
0y3yG1eA0+6Xf/d1vX4a3w3P1/A46E4/T5H3g/M5X9/c3+m//u6n+P/+X37dE
m/2f4u4n/+38U8v9yN82j/mI4p0G+R+P0w86c3mX1v8o8q6S2M/N4i1O0r1S9e
A16D/e/S9yJ2lA8S+z2T+L10Rk/xG25b8P1K3eE59M0f7x8i/9y8dG5K54C7+3
D0+g5i5024b+1q25n6c4y4P3m/t6+S3oO3yA0tQvT+A91/P4e4yv683o/1i589
eD2k5q/6/i7I+lR1g//E5vH+P8g5k7C9/f48yU53D4/4j8G+0r3x3O163xI7/
E/2uL15I76a59P+I04a43P2D60I+x0m+M0P6q53aI58cT48lR41G159E4/149
cM+28L5S664qW3n7C2vP0G5b/g759qN6c28fI2f4tXv6i5k47yW6GkX2E2pB4
A4H2a6I9fS094X/m7eC3E+0+1xY2+D4n5f8w52vO3k393E2z+7y2x6W2yL3A3f
9m2H/Ue4T3i0g63x3429+W9fXqf1a+b+f+5L0C34eE/4GvX21/xM5bS0sX1L+
l83f7+aX2n1P4M8H89A2I719aN6i1S1yO0o3e2M6wH3+r2m2t08J6T8O94b4E
9N9O5L2S2n7kX/K7K85E6r/w2yC3I+M80/k+zT1b979vM27/3N2k1/3R566C
X3rEaXoE7i/eB3M3cK1u4cE9P4C9p4/6f0T6G8y71576b9g9Pz+/o6M+lH7T9/
M8r0nI2/82k3e0+1aB5S4n3q9n29G1z+/K46O8y/e6O9yB/t0n0wO4vO195j
x9OQ0N3Y0c9f1Ld5U1vI+3fC79y7O445Xn3x3d3P3fXg+6444z3x/a+f59f6y/
p13g+80M61+L7e8+yO3L+R16E2v93X8v5C0vWv9M8n0s817u5w5e3/w88t0H69
2538O/i0B7kP9e2I9a1m3+y7x9Xv5i19b9n8i56k8eJ3L4e0y40m6O3w8q9M7/
D3m+k9m6i++3x/M2v5+n/764qf8M1+Xn/S00S0C9+t/+3O629fU821t5r1b+e
l8x9L62zJ+1yvM/T0fX9N620tC26G3u9s/33N4e1I9f4m2A6gH+cT233j5q2
bI/f0vXf84qHq0N98vX/y1/a5t85T3sS8e7A2pP85T/S+9/ZJ1l4+kE/G4J0z
x38m4J9I12O0k8Z6zN7n3eI2f2vIvf1x2vWp5yG3C+Z1y6Tj1x70vN+T44P700/
p3bO9a4n4o112sS1N/pE8v2J3R2B4/8/yQ5u7z1aJ912K+tD4x+p5e/o/4C1/I
vXvI0sJ1k+I2p+G6i+2L4P+rE/g9m+/n/O6n+/pA5L82oN/d8zN6/9xYm3T6+2
0q/c0R1q9S8a7C+j0w2oO0k2I0m1d1P4e2eX2p0+qR5i8uA+J0r0m7j5p9nI+e
/R0g8v8p/v29T/w1A7X92sN6e3E3A3qf8K3u9L3K8uC2X9p9S1S/9I930f6A6/E
o30yI4f9e1eK3z/j1k0E5/b5y4/e7g7fG6U/w8M8v/C1A/3S9C6O+Y1L2m8vO1
q//b0q3O8g6y0LqV/O7eH1e3o82sN6m+N13eA28K/p25K2a3O7M+z8599zX+t9
L3+G2c7S3q1O8Q6x5S3l7M0H593K5O/+L29X3S9s2A4C2kL0/1P/lA7I14+7a
4TfU6vj5r3z9P6+h6w3s7y93B01/w7A+041bN4P862s/i+M2m4/W8m89uG9m33
a0q+7L1tA/XvU33fMbfK389rT3C1tq/17+c28h8s8E0341l9pL0y4a3e30i1m+
5i9e/sO7g5p9i/G/2972q+/Ie79e6M3p1C6R3C1l91zT8vU6/o6z3Xf4+Pj60v
J/h9rT/49d3p/T4x6G/a3rA9p9S13tD79fB4I+iO94W2O9N4D/G094+2a3G1O/
l2L8j+9/A1a2x98O3i3x0sC2oP3/1zP8uD4o+c89uE8/E8S79/R6U1z812X452
j2i0g1yC054J4y3XyvL8p2fM26vXp9+U8D+a0n3v3Yk/C6o1y3m8L8t2I8n92l
zJvRzYy13K3l6/H30/50Ie/21l++pC1rL/E8m6r07uS5G9uE0R3JvV0C8R3wP3
X+8x7f+R2sL3i+d0j1l9uS/eJ8Ie9f3907aB3M6/qfO5vE4B/+L26X/K0x4/A/
T6G8q6/C9Y9U7W0tL8j/l3n6E2w/X5/01S19xWJ3H0g/WvF9/nS4l8R5e3xS8e
X9+35P229E85n2m/8p7N9P4y3z9u/+3tG2c78N29z8uG+5rV2j/p/+P8tD9783
43cE0K6L+S8s9fJ3L3fP2z9tUa/M/G5I3T6K0g1o2R5E6Q443r8fF7pX/0S0e2
tTveL/y7L2S9S3mJ25N0l4d+/V5z3m+K80qC+a5w3v1c+2yA1/jIuR2o8q+x/O
9y/T1p/j/+K1A/0xT3E6a7791E5+94y063e+N8q2v7W2d2j2v6S8pS21k9G2K7
i28w73eL1n/y5uC4p6pLpS7T1f6n8PZp9X+C63L2I3t0bX95iH1/tW0k1a5eIe
PzI/+b6yO6m7lA+/1mS5W4E/7/k6o9L2B/+J5k0P+7v4tE6i/G62Pj05q3A6U3
sT7G1P9C4L752mC/+nInpP2C54DvvC81S0B/u0/s4fS98fK3/49o+12y1E5k/P
q631zN/eXv1yK943a4z/a0dJ3iT0Z/n7N0a3J0o0v/yXv/k58D4/sN2B7D03I
I999D4D3p063C38xXf3I/W9O/9d2S9f2g+Q/p3A/Z2P1/fA3D1I+O0655sD8/2
Q9o5wN62R7/L1K/+hA3c6uB9k+/5G5u7qC28e+4fV6R9e70p273v2Ld++2A79v
T877h+9T3x25O5fA55n9o1jL7E9a5W8m2G6S1G0m//0p/2X1/2j2v/4B4/A/M/
u5s/2f8r8v8q4L5f08l34/c2+7uW7d/aG3O4f8pI8S+yO91f/k4v+h+/l/eB1S
4c/l93K/+a0O4H6H0b35K2Pj4y312lE14j+C4J5t8A0R5u4H4u2B/tN3hXz0eC
0t5x49Pse/5E0P/tD//920S/f8uT9S3x/Y+2f+2vJ+Q622R8xL13v3o3+3c6uB
T96uL955f2+1R3J/vL3G21S13q7D0lMIn1+d7aA6/c4t3E1r06f1sH6U2R96fJ
8k8P0tO3g852o3eC+6Xn3q3IeM9S3O8P6G9fV7G3A6I/o1m0M6/sR+d0uK5A19
/I2z8sT/4/19g3271vL8z5E3/u03O5N+/mN9/c+S+733e3B2M84Tze2D1X9v6A
4T/6jVfV6j1PqH+z4uJ8f36C9L9M8a59J3c/V/D1P6G+/5o01A7k923N+/B6J5
yD8r+26u9yP9N20G/eL+s3E/eG0m72o60A/G43zM8/fD/eA0m5p24L26E0lD8T
Xp7f/xI0C//vA1s75o35/49uT+K+5l35L5r9O0q8H+b875H+I98eUa4L86sL9e
A6x+O12C8O4J2B0X4I4J7v/n1JvF+w5k3h++F2nS0l6uFvC7L0eS435+80T2x6
71+X5i22uA+sK5W1k14v6j5o8N025w34l++x1/3c2E098d6E/1p/pC1+fP/p44
C/+N8J6V0n+0h45XoDvf4tL68dD/C8r766/i8S2Gf37uX53x2s9m//S3d8R453
+A05+/r7y7x5xXbU35p4+S/y+86q447u/1k4C1S4uO+G4A/+w22S5f43A1n3/a
1z2+v2A3yI2y8O4i3N4j67u03C3d/+y69739P4P7m8A2461/c2O93M3m3P5p4x
B//5c7/z7uE+8q1B2H6XInzOq4uQ7n5n7d//f3uT3X+w3t6xP6j8t4v9X+D6C4
r93f1j1s9n9a/aC2X2t0B/eU9q/i8L2k3v5B4o7C8eA606f3o2d+k+2A9e+P6r
Wz6a3n8y96S44//63D5Xo//7/d57I1kX3s0A9tD980C5k7l02/33rG9E8oD/I/
g9M8L/dM+s2a24y/+iX6fH8P2uL6i1P/+eG25uE/U/s3D4j5v3w3B/e/E0mO28
L5C4dO7L6M//e4L9v/6eE2lO6X7z9X7P6q3g0tM//T553M9L6I5A+G2e1C4X0j
iP896zC08N3kIn1t6/uO/M8963+4x3uE/Zf4O++4m9u2z+U/lveU5M9x3a0C+S
8E0zI/+N+7/1+j2G6m47u/uG0e4k3b9Xf1/K2j+6m/d+o/I+X3t7P/u/2D5i56
C5i66x6r43/+yv9/+p6+D8x/+sHn8m0v1r2C2X4S63C++d5I//K5z66R+xI5
T/0K/p0g//P33A2O3b7n359o03n+1bU/O//L+f4/3K4L2uNf35B/+A5/s/q
C0L8k+m///O6C39k/72U5r+K5J64//n43eB0t7z2d3T/v9fU8217Xv0L+/0x
1w/J3Y3BfQ///L83v+8d5v29f0JzC/A8i5n+i2g7W23L2v1zH+/tX/v3N8+
9N/4y2L++4x18D27//3L5b/g7/tD/f+P6wP+/L2m43aI+7S8G48vC+B5P//uA3
E7P4R8sO58fJ3q38rA/+L1z3739/W/+/O9n7y/k380R8h2C7C43+E4nO+M9S0
7sI1f/S8b2/6P27/zP9p+k5q/62X/s1N3+lI0sD3q1s+p+/1tH7c5+R/3p
I8D2a3O77nC6s9j7O/B8f+i/9C9tA6y5/0vO3D2K5H98v/7E0D3rI1619T/u0e
f5G/+t7n//c9X//K/+x/9kL4rI/+T/+a1r+P9a1eNf13k4i5o7/x2C1sL6C1u9
r0v/o2m9a/S/N4P6S+R9U9vXz7w1D7R9C9O367m++u8C5q3o7r2X2D3O6W7K4u
E+sA3fX30x0J5S3nL/D092523v1y2A/+j/uE37G5e/C+B7+X1eN4O4z2a+/iP2
9R7M2f3n1X2K/a+r3a9m9939L7y/b1a//D8o8N/M2i7x18v/G//s+J369J3
lH9X/x3296/G6I03T5P12uG6u5N13c8r127m54z+T/sXvR2D5N11q4k//S3w2/
E1q063uC092q/+uE928Gv93T9m///8S4x/t4E11+B/51P+2M0vD6n9O3sO6f1s
P2e2N4t+/rE64S9w4i4D5/b0uD9n2z3E9b9G5/M//U1e6A0T+3+O/+d2zE/M/+
b8r+C/2pM9P6I3e22p+w/vD4O8A1f30I7W0a7H+a5uA7D3H/5uI80rA3j++r//
8v5u5s/m/b1G4e/v4T40eC1A/3S/fN/m5D/65f0u//T4zI6I9c14+L3O9P
3d6S2s/x/+p//fC2s++
"""

st.title("📄 심층면담 회의록 문서 & 로그 자동 생성기")
st.markdown("🔗 연동 시트: **[구글 스프레드시트 (심층면담_회의록)](https://docs.google.com/spreadsheets/d/1ws9JTAdRXwbp--NhrjWwelNorSTv1_LIJW7DijUtJLU/edit?gid=770556375#gid=770556375)**")
st.caption("구글 시트의 최신 데이터를 실시간으로 가져와 HWPX 회의록과 엑셀 점검 로그를 일괄 생성합니다.")

st.divider()

# 구글 시트 정보
SHEET_ID = "1ws9JTAdRXwbp--NhrjWwelNorSTv1_LIJW7DijUtJLU"
GID = "770556375" # '심층면담_회의록' 시트 GID
GOOGLE_SHEET_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=csv&gid={GID}"

# 메인 버튼 하나로 동작!
if st.button("🚀 실시간 데이터 읽기 및 회의록 자동 생성 시작", type="primary", use_container_width=True):
    
    with st.spinner("🔄 구글 시트에서 최신 데이터를 불러오는 중..."):
        try:
            df_raw = pd.read_csv(GOOGLE_SHEET_URL)
            data_df = df_raw.iloc[2:].dropna(subset=['문서ID']).copy()
        except Exception as e:
            st.error(f"❌ 구글 시트 데이터를 읽어오는 중 오류가 발생했습니다: {e}")
            st.stop()
            
    st.success(f"✅ 구글 시트 연결 성공! 총 **{len(data_df)}건**의 회의록 데이터를 불러왔습니다.")

    # 템플릿 로드 (내장된 Base64 디코딩)
    hwpx_bytes = base64.b64decode(TEMPLATE_BASE64.strip())
    hwpx_zip = zipfile.ZipFile(io.BytesIO(hwpx_bytes), 'r')
    template_files = {name: hwpx_zip.read(name) for name in hwpx_zip.namelist()}
    
    sec0_text = template_files['Contents/section0.xml'].decode('utf-8')
    pattern = r'<hp:stringParam name="Command">(.*?)</hp:stringParam>'
    commands = re.findall(pattern, sec0_text)
    
    merge_fields = []
    for cmd in commands:
        if cmd not in merge_fields:
            merge_fields.append(cmd)
            
    total_merge_cnt = len(merge_fields) # 14개

    # 회의일시 순서로 정렬
    data_df['일시_dt'] = pd.to_datetime(data_df['일시'], errors='coerce')
    sorted_df = data_df.sort_values(by=['일시_dt', '시작시간', '학교명'], na_position='last').reset_index(drop=True)

    version_str = f"v.{datetime.datetime.now().strftime('%y%m%d_%H%M')}"
    folder_name = f"결과물_심층면담_회의록_{version_str}"

    # 메모리 ZIP 생성
    zip_buffer = io.BytesIO()
    log_records = []
    
    progress_bar = st.progress(0)
    progress_text = st.empty()

    with zipfile.ZipFile(zip_buffer, 'w', compression=zipfile.ZIP_DEFLATED) as main_zip:
        total_rows = len(sorted_df)
        
        for idx, (_, row) in enumerate(sorted_df.iterrows(), start=1):
            doc_id = str(row['문서ID']).strip()
            school = str(row['학교명']).strip() if pd.notna(row['학교명']) else ""
            round_num = str(row['회차']).strip() if pd.notna(row['회차']) else ""
            
            date_val = row['일시']
            if pd.isna(date_val):
                date_str = ""
            elif isinstance(date_val, (pd.Timestamp, datetime.datetime)):
                date_str = date_val.strftime('%Y-%m-%d')
            else:
                date_str = str(date_val).strip()
                
            xml_content = template_files['Contents/section0.xml'].decode('utf-8')
            missing_fields = []
            
            for col in data_df.columns:
                if col == '일시_dt':
                    continue
                val = row[col]
                
                if pd.isna(val) or str(val).strip() == "" or str(val).strip().lower() == "nan":
                    val_str = ""
                    if col in merge_fields:
                        missing_fields.append(col)
                elif isinstance(val, (pd.Timestamp, datetime.datetime)):
                    val_str = val.strftime('%Y-%m-%d')
                elif isinstance(val, datetime.time):
                    val_str = val.strftime('%H:%M')
                else:
                    val_str = str(val).strip()
                    
                val_str = val_str.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
                xml_content = xml_content.replace(f"{{{{{col}}}}}", val_str)
                
            # HWPX 메모리 생성
            doc_buffer = io.BytesIO()
            with zipfile.ZipFile(doc_buffer, 'w', compression=zipfile.ZIP_DEFLATED) as z_out:
                for name, data in template_files.items():
                    if name == 'Contents/section0.xml':
                        z_out.writestr(name, xml_content.encode('utf-8'))
                    else:
                        z_out.writestr(name, data)
                        
            main_zip.writestr(f"{folder_name}/{doc_id}.hwpx", doc_buffer.getvalue())
            
            missing_cnt = len(missing_fields)
            ratio_str = f"{missing_cnt}건 / {total_merge_cnt}건"
            status = "정상" if missing_cnt == 0 else "일부항목누락"
            
            row_dict = {
                "생성번호": idx,
                "문서ID": doc_id,
                "학교명": school,
                "회의일시": date_str,
                "회차": round_num,
                "생성상태": status,
                "누락현황(누락/전체)": ratio_str
            }
            
            for field in merge_fields:
                if field in missing_fields:
                    row_dict[f"[점검]{field}"] = f"{field} N/A"
                else:
                    row_dict[f"[점검]{field}"] = "-"
                    
            log_records.append(row_dict)
            
            progress_bar.progress(idx / total_rows)
            progress_text.text(f"⚡ HWPX 회의록 생성 중... [{idx}/{total_rows}] {doc_id}.hwpx")

        log_df = pd.DataFrame(log_records)
        
        # 엑셀 로그 파일 작성 (빨간 글씨 서식 적용)
        excel_buffer = io.BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "생성로그"
        
        headers = list(log_df.columns)
        ws.append(headers)
        
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        red_text_font = Font(color="C00000", bold=True)
        warning_text_font = Font(color="9C6500", bold=True)
        
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        for r_idx, row_data in enumerate(log_records, start=2):
            for c_idx, col_name in enumerate(headers, start=1):
                val = row_data[col_name]
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.border = thin_border
                
                if col_name in ["생성번호", "회의일시", "회차", "생성상태", "누락현황(누락/전체)"]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                if col_name.startswith("[점검]") and str(val).endswith("N/A"):
                    cell.font = red_text_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col_name == "생성상태" and val == "일부항목누락":
                    cell.font = warning_text_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    
        wb.save(excel_buffer)
        
        # ZIP 내에 로그 파일 2종 포함
        excel_filename = f"심층면담_서류_Log_{version_str}.xlsx"
        csv_filename = f"심층면담_서류_Log_{version_str}.csv"
        
        main_zip.writestr(f"{folder_name}/{excel_filename}", excel_buffer.getvalue())
        main_zip.writestr(f"{folder_name}/{csv_filename}", log_df.to_csv(index=False, encoding='utf-8-sig'))

    st.success("🎉 회의록 58개 및 점검 로그 완성이 완료되었습니다!")
    
    # 압축 파일 1클릭 다운로드
    st.download_button(
        label=f"📦 {folder_name}.zip 압축 파일 다운로드",
        data=zip_buffer.getvalue(),
        file_name=f"{folder_name}.zip",
        mime="application/zip",
        use_container_width=True
    )
    
    with st.expander("📋 생성 로그 데이터 미리보기"):
        st.dataframe(log_df)